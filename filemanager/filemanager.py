import openpyxl
import os
from datetime import datetime
import csv
from openpyxl.styles import Alignment

FREE_MARK = ' '

def read_csv_settings(file_path) -> dict:
    settings = {}
    with open(file_path, 'r', newline='') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=";")
        settings = {k:[v] for k, v in next(reader).items()}
            
        for row in reader:
            for header in settings:
                if row[header]:
                    settings[header] += [row[header]]

    # change {key : [one value]} to {key : one value}
    for header in settings:
        if len(settings[header]) <= 1:
            settings[header] = settings[header][0]       
    
    return settings

def save_orders_to_file(orders, file_path, save_backup=False):        
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row in orders:
        sheet.append(row)

    adjust_sheet_cells(sheet=sheet)
    workbook.save(filename=file_path)

    if save_backup == "Yes":
        date = datetime.today().strftime(r"%H.%M %d-%m-%Y")
        dir_path = os.path.dirname(file_path)
        workbook.save(filename=os.path.join(dir_path, 'Backups', f'{date}.xlsx'))

def save_employers_data_to_table(employers_data : dict, table_path):
    workbook = openpyxl.load_workbook(table_path)
    if not 'Employers' in workbook.sheetnames:
        workbook.create_sheet(title='Employers')
    sheet = workbook['Employers']

    for col, name in enumerate(employers_data):
        sheet.cell(row=1, column=col * 2 + 1, value=name)

        orders = employers_data[name]
        for row in range(len(orders)):
            sheet.cell(row=row + 2, column=col * 2 + 1, value=orders[row][0])
            sheet.cell(row=row + 2, column=col * 2 + 2, value=orders[row][1])

    adjust_sheet_cells(sheet=sheet)
    
    for col in range(len(employers_data)):
        sheet.merge_cells(start_row=1, start_column=col * 2 + 1, end_row=1, end_column=col * 2 + 2)
        sheet.cell(row=1, column=col * 2 + 1).alignment = Alignment(horizontal='center')
        
    workbook.save(table_path)

def get_employers_from_txt(file_path) -> list:
    with open(file_path, 'r') as f:
        return [line.strip() for line in f.readlines()]

def get_orders_from_file(file_path):
    workbook = openpyxl.load_workbook(filename=file_path)
    sheet = workbook.active
    orders = []

    for row in range(len(list(sheet.rows))):
        order = []
        for col in range(1, 5):
            cell = sheet.cell(row=row + 1, column=col)
            order.append(cell.value if cell.value else FREE_MARK)

        orders.append(order)
    
    return orders

def adjust_sheet_cells(sheet):
    for col in sheet.columns:
        max_width = 10
        for cell in col:
            if len(str(cell.value)) > max_width:
                max_width = len(str(cell.value))

        sheet.column_dimensions[col[0].column_letter].width = max_width * 1.2

if __name__ == "__main__":
    print(read_csv_settings('D:\Програмування\Projects\Suplier Manager\settings\settings.csv'))
